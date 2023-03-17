import mne as mne
import numpy as np
import pandas as pd
import torch as torch
import time
from tqdm import tqdm
from openpyxl import load_workbook

from preprocessing import Preprocessing

import torch
import torch.nn as nn
 
from subDataset import subDataset
import torch.utils.data.dataloader as DataLoader

def accuracy(predictions,labels):
    criterion1 = torch.nn.MSELoss()
    MSE = criterion1(predictions, labels)
    return MSE,len(labels)

class topo_CNN(nn.Module):
    def __init__(self, topo_map):
        super(topo_CNN,self).__init__()
        self.topo_map = topo_map
        self.attention_net = []
        for i in range(64):
            self.attention_net.append(nn.Sequential(
            nn.Linear(64, 128, bias=False),
            nn.ReLU(inplace=True),
            nn.Linear(128, 64, bias=False),
            nn.Sigmoid()
        ).cuda())
        self.fc = nn.Sequential(
            nn.Linear(64, 16, bias=False),
            nn.ReLU(inplace=True),
            nn.Linear(16, 64, bias=False),
            nn.Sigmoid()
        )
        self.conv1= nn.Sequential(
            nn.Conv1d(
                in_channels=64,
                out_channels=32,
                kernel_size=5,
                stride=1,
                padding=2,
            ),
            nn.ReLU(),
            nn.MaxPool1d(kernel_size=2),
        )
        self.conv2= nn.Sequential(
            nn.Conv1d(
                in_channels=32,
                out_channels=16,
                kernel_size=5,
                stride=1,
                padding=2,
            ),
            nn.ReLU(),
            nn.MaxPool1d(kernel_size=2),
        )
        self.conv3= nn.Sequential(
            nn.Conv1d(
                in_channels=16,
                out_channels=8,
                kernel_size=5,
                stride=1,
                padding=2,
            ),
            nn.ReLU(),
            nn.MaxPool1d(kernel_size=2),
        )
        self.conv4= nn.Sequential(
            nn.Conv1d(
                in_channels=8,
                out_channels=4,
                kernel_size=5,
                stride=1,
                padding=2,
            ),
            nn.ReLU(),
            nn.MaxPool1d(kernel_size=2),
        )
        self.conv5= nn.Sequential(
            nn.Conv1d(
                in_channels=4,
                out_channels=2,
                kernel_size=5,
                stride=1,
                padding=2,
            ),
            nn.ReLU(),
            nn.MaxPool1d(kernel_size=2),
        )
        self.conv6= nn.Sequential(
            nn.Conv1d(
                in_channels=2,
                out_channels=1,
                kernel_size=5,
                stride=1,
                padding=2,
            ),
            nn.ReLU(),
            nn.MaxPool1d(kernel_size=2),
        )
        self.out=nn.Linear(16,1)

    def forward(self,x):
        # attention = self.avg_pool(x)
        # attention = self.fc(attention.reshape(1, 64)).unsqueeze(2).repeat(1, 1, 1024)
        xmean = x[0].mean(axis=1).reshape((1,64))
        att = xmean * self.topo_map
        attention = []
        for i in range(64):
            net = self.attention_net[i]
            attention.append(net(att[i]))
        attentions = torch.tensor([item.cpu().detach().numpy() for item in attention]).cuda()
        attentions = attentions.reshape(1, 64, 64)
        x = attentions.matmul(x)
        x= self.conv1(x)
        x= self.conv2(x)
        x= self.conv3(x)
        x= self.conv4(x)
        x= self.conv5(x)
        x= self.conv6(x)
        output = self.out(x)
        return output.reshape((1, 1))

def run(train_loader, test_loader):
    wb = load_workbook('topo.xlsx')
    sheets = wb.worksheets

    sheet1 = sheets[0]
    topo = []
    for i in range(64):
        line = []
        for j in range(64):
            line.append(sheet1.cell(i + 3, j + 3).value)
        topo.append(line)

    topo_map = torch.tensor(topo).to(torch.float32).cuda()
    
    net=topo_CNN(topo_map)
    net = net.cuda()

    critertion=nn.MSELoss()

    optimizer =torch.optim.Adam(net.parameters(),lr=0.001)
    
    num_epoche = 100
    loss_list = []
    mse_list = []
    for epoch in range(num_epoche):
        train_rights=0.
        losses = 0.
        pbar = tqdm(train_loader, total=30, leave=True, ncols=80)
        pbar.set_description(f'epoch {(epoch + 1)}/{num_epoche}')
        for batch_idx,(data,target) in enumerate(pbar):
            data = data.cuda()
            target = target.cuda()
            net.train()
            output = net(data)
            loss = critertion(output, target)
            optimizer.zero_grad()
            loss.to(torch.float32)
            losses += loss.item()
            loss.backward()
            optimizer.step()

            pbar.set_postfix(loss = losses / (batch_idx + 1))

            if (batch_idx + 1) % 30 == 0:
                mses = 0.
                for batch_id, (x, y) in enumerate(test_loader):
                    x = x.cuda()
                    y = y.cuda()
                    pred = net(x)
                    mse, length = accuracy(pred, y)
                    mses += mse.item()
                pbar.set_postfix(loss = losses / (batch_idx + 1), mse = mses / 20)
                if ((epoch + 1) % 5 == 0) or (epoch == 0):
                    mse_list.append(mses / 20)
                    loss_list.append(losses / (batch_idx + 1))
                    
    torch.save(net, 'model.pkl')
    return loss_list, mse_list